from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
<<<<<<< HEAD
from django.http import HttpResponse
from django.core.paginator import Paginator
from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem
import io
import openpyxl
from django.core.mail import EmailMessage


# Главная страница
def index(request):
    """Главная страница с популярными товарами и категориями"""
    popular_products = Product.objects.order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'popular_products': popular_products,
        'categories': categories,
    })


# 1. Список товаров (Каталог) с пагинацией
def product_list(request):
    """Страница каталога товаров с фильтрацией и пагинацией"""
=======
from .models import Product, Category, Manufacturer, Cart, CartItem

# 1. Список товаров (Каталог)
def product_list(request):
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    products = Product.objects.all()
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

<<<<<<< HEAD
    # Поиск
=======
    # Поиск по названию ИЛИ описанию (используем Q-объекты)
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    search_query = request.GET.get('search', '')
    if search_query:
        products = products.filter(
            Q(title__icontains=search_query) | Q(description__icontains=search_query)
        )

<<<<<<< HEAD
    # Фильтр по категории
=======
    # Фильтрация по категории
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    category_id = request.GET.get('category', '')
    if category_id:
        products = products.filter(category_id=category_id)

<<<<<<< HEAD
    # Фильтр по производителю
=======
    # Фильтрация по производителю
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    manufacturer_id = request.GET.get('manufacturer', '')
    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

<<<<<<< HEAD
    # Пагинация: 9 товаров на странице
    paginator = Paginator(products, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'shop/catalog.html', {
        'products': page_obj,
=======
    return render(request, 'shop/product_list.html', {
        'products': products,
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        'categories': categories,
        'manufacturers': manufacturers,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_manufacturer': manufacturer_id,
<<<<<<< HEAD
        'page_obj': page_obj,
    })


=======
    })

>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
# 2. Детальная информация о товаре
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})

<<<<<<< HEAD

=======
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
# 3. Добавление товара в корзину
@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_item, item_created = CartItem.objects.get_or_create(cart=cart, product=product)
    
    if not item_created:
        if cart_item.quantity + 1 <= product.stock_quantity:
            cart_item.quantity += 1
            cart_item.save()
            messages.success(request, f"Количество товара {product.title} увеличено.")
        else:
            messages.error(request, "Недостаточно товара на складе.")
    else:
        messages.success(request, f"Товар {product.title} добавлен в корзину.")
        
<<<<<<< HEAD
    return redirect('store:product_list')

=======
    return redirect('product_list')
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c

# 4. Обновление количества товара в корзине
@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    if request.method == 'POST':
        new_quantity = int(request.POST.get('quantity', 1))
        
        if new_quantity <= cart_item.product.stock_quantity:
            cart_item.quantity = new_quantity
            cart_item.save()
            messages.success(request, "Количество обновлено.")
        else:
            messages.error(request, f"На складе доступно только {cart_item.product.stock_quantity} шт.")
            
<<<<<<< HEAD
    return redirect('store:cart_view')

=======
    return redirect('cart_view')
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c

# 5. Удаление товара из корзины
@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    messages.success(request, "Товар удален из корзины.")
<<<<<<< HEAD
    return redirect('store:cart_view')

=======
    return redirect('cart_view')
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c

# 6. Просмотр корзины
@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    return render(request, 'shop/cart.html', {'cart': cart})
<<<<<<< HEAD

=======
import io
import openpyxl
from django.core.mail import EmailMessage
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c

@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    cart_items = cart.items.all()

<<<<<<< HEAD
    if not cart_items:
        messages.error(request, "Ваша корзина пуста. Оформление невозможно.")
        return redirect('store:cart_view')

    if request.method != 'POST':
        return render(request, 'shop/checkout.html')

    user_address = request.POST.get('address', '')
    user_comment = request.POST.get('comment', 'Без комментария')

=======
    # Если в корзине пусто, не пускаем дальше
    if not cart_items:
        messages.error(request, "Ваша корзина пуста. Оформление невозможно.")
        return redirect('cart_view')

    # Если пользователь зашел первый раз — открываем форму ввода адреса
    if request.method != 'POST':
        return render(request, 'shop/checkout.html')

    # Если пользователь отправил форму (POST-запрос):
    # Принимаем данные, введенные пользователем
    user_address = request.POST.get('address', '')
    user_comment = request.POST.get('comment', 'Без комментария')

    # Создаем Excel книгу в памяти для чека
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Чек заказа"

<<<<<<< HEAD
=======
    # Заполняем накладную текстовыми пояснениями
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    ws['A1'] = "ЭЛЕКТРОННЫЙ ЧЕК СПОРТИВНОГО МАГАЗИНА"
    ws['A2'] = f"Покупатель (User): {request.user.username}"
    ws['A3'] = f"Адрес доставки: {user_address}"
    ws['A4'] = f"Комментарий: {user_comment}"
    ws['A5'] = "--------------------------------------------------"

<<<<<<< HEAD
=======
    # Шапка таблицы
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    ws.append(["Наименование товара", "Количество", "Цена за шт.", "Итоговая стоимость"])

    total_price = 0
    for item in cart_items:
        item_total = item.product.price * item.quantity
        total_price += item_total
<<<<<<< HEAD
        ws.append([item.product.title, item.quantity, float(item.product.price), float(item_total)])
=======

        # Добавляем строку товара в Excel
        ws.append([item.product.title, item.quantity, float(item.product.price), float(item_total)])

        # Списываем купленное количество со склада
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        item.product.stock_quantity -= item.quantity
        item.product.save()

    ws.append([])
    ws.append(["ОБЩАЯ СУММА К ОПЛАТЕ:", "", "", float(total_price)])

<<<<<<< HEAD
=======
    # Сохраняем сгенерированный Excel в байтовый поток памяти для отправки по почте
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

<<<<<<< HEAD
=======
    # Формируем и отправляем Email с вложением чека (Задание 4)
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
    user_email = request.user.email if request.user.email else f"{request.user.username}@example.com"

    subject = f"Ваш чек заказа в магазине Спортивных товаров #{cart.id}"
    body = (
        f"Здравствуйте, {request.user.username}!\n\n"
<<<<<<< HEAD
        f"Ваш заказ успешно сформирован.\n"
        f"Адрес доставки: {user_address}\n"
        f"Комментарий: {user_comment}\n"
        f"Итого к оплате: {total_price} руб.\n\n"
        f"Чек во вложении (Excel)."
    )

    email = EmailMessage(subject, body, 'shop@sport-accessories.com', [user_email])
    email.attach('sport_order_check.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()

    cart_items.delete()

    return HttpResponse(
        f"<body style='font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; max-width: 600px; background: #fafafa;'>"
        f"<div style='background: #fff; padding: 30px; border: 2px solid #28a745; border-radius: 8px;'>"
        f"<h2 style='color: #28a745;'>🎉 Заказ успешно завершен!</h2>"
        f"<p>Чек отправлен на: <u>{user_email}</u></p>"
        f"<p>Адрес доставки: {user_address}</p>"
        f"<p>Итого: <b style='color: #e44d26;'>{total_price} руб.</b></p>"
        f"<a href='/catalog/' style='display: inline-block; background: #007bff; color: white; padding: 10px 15px; text-decoration: none; border-radius: 4px;'>Вернуться в каталог</a>"
        f"</div></body>"
    )


# API ViewSets
from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
=======
        f"Ваш заказ успешно сформирован и передан в курьерскую службу.\n"
        f"Пояснение к вашим данным:\n"
        f"- Адрес доставки: {user_address}\n"
        f"- Комментарий: {user_comment}\n"
        f"- Итого к оплате: {total_price} руб.\n\n"
        f"Подробный электронный чек находится во вложении (формат Excel).\n"
        f"Спасибо за покупку!"
    )

    email = EmailMessage(
        subject,
        body,
        'shop@sport-accessories.com',
        [user_email]
    )

    # Прикрепляем созданный Excel файл к письму
    email.attach('sport_order_check.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    email.send()

    # Очищаем корзину пользователя после успешного оформления
    cart_items.delete()

    # Выводим результаты обработки на экран с подробными пояснениями для ЛР
    return HttpResponse(
        f"<body style='font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; max-width: 600px; background: #fafafa;'>"
        f"<div style='background: #fff; padding: 30px; border: 2px solid #28a745; border-radius: 8px; box-shadow: 0 4px 6px rgba(0,0,0,0.05);'>"
        f"<h2 style='color: #28a745; margin-top: 0;'>🎉 Заказ успешно завершен!</h2>"
        f"<p><b>Пояснение выводимых результатов обработки вашей формы:</b></p>"
        f"<ul>"
        f"<li><b>Статус отправки:</b> Чек успешно сформирован в Excel и отправлен на ваш email: <u>{user_email}</u></li>"
        f"<li><b>Указанный адрес доставки:</b> {user_address}</li>"
        f"<li><b>Ваш комментарий:</b> {user_comment}</li>"
        f"<li><b>Итоговая стоимость:</b> <b style='color: #e44d26;'>{total_price} руб.</b></li>"
        f"<li><b>Состояние корзины:</b> Корзина полностью очищена.</li>"
        f"</ul>"
        f"<hr style='border: 0; border-top: 1px solid #eee; margin: 20px 0 Triton;'>"
        f"<a href='/catalog/' style='display: inline-block; background: #007bff; color: white; padding: 10px 15px; text-decoration: none; border-radius: 4px; font-weight: bold;'>Вернуться в каталог товаров</a>"
        f"</div>"
        f"</body>"
    )
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Manufacturer, Category, Product, Cart, CartItem, Order, OrderItem
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
from .serializers import (
    ManufacturerSerializer, CategorySerializer, ProductSerializer,
    CartSerializer, CartItemSerializer, OrderSerializer, OrderItemSerializer
)

<<<<<<< HEAD
=======

# API для Производителей
class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


# API для Категорий
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]


# API для Товаров
from rest_framework import viewsets, permissions, filters
from django_filters.rest_framework import DjangoFilterBackend
from .models import Manufacturer, Category, Product, Cart, CartItem, Order, OrderItem
from .serializers import (
    ManufacturerSerializer, CategorySerializer, ProductSerializer,
    CartSerializer, CartItemSerializer, OrderSerializer, OrderItemSerializer
)


>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

<<<<<<< HEAD
=======

>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]

<<<<<<< HEAD
=======

>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.IsAuthenticatedOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'manufacturer']
    search_fields = ['title', 'description']
    ordering_fields = ['price', 'stock_quantity', 'title']
    ordering = ['title']

<<<<<<< HEAD
=======

class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]


class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated]

# API для Корзины
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
<<<<<<< HEAD
=======
        # Пользователь видит только свою корзину
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        if self.request.user.is_authenticated:
            return Cart.objects.filter(user=self.request.user)
        return Cart.objects.none()

    def perform_create(self, serializer):
<<<<<<< HEAD
=======
        # Автоматически привязываем корзину к текущему пользователю
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'])
    def add_item(self, request, pk=None):
<<<<<<< HEAD
=======
        """Добавить товар в корзину"""
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        cart = self.get_object()
        product_id = request.data.get('product')
        quantity = request.data.get('quantity', 1)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response({'error': 'Товар не найден'}, status=status.HTTP_404_NOT_FOUND)
        
        cart_item, created = CartItem.objects.get_or_create(
            cart=cart,
            product=product,
            defaults={'quantity': quantity}
        )
        
        if not created:
            cart_item.quantity += int(quantity)
            cart_item.save()
        
        serializer = CartItemSerializer(cart_item)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

<<<<<<< HEAD
=======

# API для Элементов корзины
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
<<<<<<< HEAD
        cart, _ = Cart.objects.get_or_create(user=self.request.user)
        serializer.save(cart=cart)

=======
        # Автоматически привязываем к корзине текущего пользователя
        cart, _ = Cart.objects.get_or_create(user=self.request.user)
        serializer.save(cart=cart)


# API для Заказов
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
<<<<<<< HEAD
=======
        # Пользователь видит только свои заказы
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        if self.request.user.is_authenticated:
            return Order.objects.filter(user=self.request.user)
        return Order.objects.none()

    def perform_create(self, serializer):
<<<<<<< HEAD
=======
        # Создаем заказ из корзины
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        cart = Cart.objects.filter(user=self.request.user).first()
        if not cart or not cart.items.exists():
            raise Exception("Корзина пуста")
        
        order = serializer.save(user=self.request.user, total_price=cart.total_price)
        
<<<<<<< HEAD
=======
        # Переносим товары из корзины в заказ
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        for cart_item in cart.items.all():
            OrderItem.objects.create(
                order=order,
                product=cart_item.product,
                quantity=cart_item.quantity,
                price=cart_item.product.price
            )
        
<<<<<<< HEAD
=======
        # Очищаем корзину
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        cart.items.all().delete()

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
<<<<<<< HEAD
=======
        """Обновить статус заказа (только для админов)"""
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
        if not request.user.is_staff:
            return Response({'error': 'Только администраторы могут менять статус'}, 
                          status=status.HTTP_403_FORBIDDEN)
        
        order = self.get_object()
        new_status = request.data.get('status')
        
        if new_status not in dict(Order.STATUS_CHOICES):
            return Response({'error': 'Неверный статус'}, status=status.HTTP_400_BAD_REQUEST)
        
        order.status = new_status
        order.save()
        
        serializer = self.get_serializer(order)
        return Response(serializer.data)

<<<<<<< HEAD
=======

# API для Элементов заказа
>>>>>>> d84a7fc3a4c40612c8e4545f3018aa927d65582c
class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.all()
    serializer_class = OrderItemSerializer
    permission_classes = [permissions.IsAuthenticated]